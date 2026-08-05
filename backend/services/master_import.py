"""Import the DB1 master tag registry from the lab Excel workbook.

The master registry is intentionally one row per unique EPC. It should not
contain scan events; DB2 remains the place for incoming scan history.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from backend import models


FIELD_MAP = {
    "Record_ID": "record_id",
    "Asset_Name": "asset_name",
    "EPC": "epc",
    "Tag_ID": "tag_id",
    "Component_Location": "component_location",
    "Zone_Code": "zone_code",
    "Zone_Name": "zone_name",
    "Sub_Zone": "sub_zone",
    "Category": "category",
    "Manufacturer": "manufacturer",
    "Model": "model",
    "Unit": "unit",
    "Qty": "qty",
    "Status": "status",
    "Reader_ID": "reader_id",
    "Reader_Type": "reader_type",
    "Verification_Method": "verification_method",
    "Manual_Check": "manual_check",
    "Comments": "comments",
}


@dataclass
class MasterImportReport:
    source: str
    sheet: str = "DB1_MASTER"
    rows_seen: int = 0
    unique_records: int = 0
    inserted: int = 0
    updated: int = 0
    deleted: int = 0
    skipped: int = 0
    duplicate_epcs: list[dict[str, Any]] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def import_master_from_excel(
    db: Session,
    workbook_path: str | Path,
    *,
    replace: bool = False,
    sheet_name: str = "DB1_MASTER",
) -> MasterImportReport:
    """Import unique DB1 master tags from an Excel workbook.

    Args:
        db: SQLAlchemy session.
        workbook_path: Path to the lab workbook.
        replace: When true, remove DB1 rows that are not in the workbook.
        sheet_name: Worksheet containing DB1 master rows.

    Duplicate EPCs are skipped after the first occurrence because DB1 is the
    unique tag registry. The report tells the operator which rows need cleanup.
    """

    path = Path(workbook_path)
    report = MasterImportReport(source=str(path), sheet=sheet_name)
    if not path.exists():
        report.errors.append(f"Workbook not found: {path}")
        return report

    try:
        import openpyxl
    except ImportError:
        report.errors.append("openpyxl is required to import .xlsx files")
        return report

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        report.errors.append(f"Sheet {sheet_name!r} not found")
        return report

    worksheet = workbook[sheet_name]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in header_row]

    records: list[dict[str, Any]] = []
    seen_epcs: dict[str, str] = {}
    seen_record_ids: set[str] = set()

    for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        raw = dict(zip(headers, values))
        if not raw.get("Record_ID") and not raw.get("EPC"):
            continue
        report.rows_seen += 1

        record = _normalize_master_row(raw)
        error = _validate_record(record, row_number)
        if error:
            report.errors.append(error)
            report.skipped += 1
            continue

        epc = record["epc"]
        if epc in seen_epcs:
            report.duplicate_epcs.append({
                "epc": epc,
                "kept_record_id": seen_epcs[epc],
                "skipped_record_id": record["record_id"],
                "row": row_number,
            })
            report.skipped += 1
            continue

        if record["record_id"] in seen_record_ids:
            report.errors.append(f"Duplicate Record_ID {record['record_id']} at Excel row {row_number}")
            report.skipped += 1
            continue

        seen_epcs[epc] = record["record_id"]
        seen_record_ids.add(record["record_id"])
        records.append(record)

    report.unique_records = len(records)

    source_record_ids = {record["record_id"] for record in records}
    if replace:
        stale = db.query(models.Asset).filter(~models.Asset.record_id.in_(source_record_ids)).all()
        report.deleted = len(stale)
        for asset in stale:
            db.delete(asset)
        db.flush()

    for record in records:
        asset = db.query(models.Asset).filter(models.Asset.record_id == record["record_id"]).first()
        if not asset:
            asset = db.query(models.Asset).filter(models.Asset.epc == record["epc"]).first()

        if asset:
            for key, value in record.items():
                setattr(asset, key, value)
            report.updated += 1
        else:
            db.add(models.Asset(**record))
            report.inserted += 1

    db.commit()
    return report


def _normalize_master_row(raw: dict[str, Any]) -> dict[str, Any]:
    record: dict[str, Any] = {}
    for excel_name, model_name in FIELD_MAP.items():
        record[model_name] = _clean(raw.get(excel_name))

    record["record_id"] = str(record["record_id"]).upper()
    record["epc"] = str(record["epc"]).upper()
    record["zone_code"] = str(record["zone_code"]).upper() if record.get("zone_code") else None
    record["qty"] = _to_int(record.get("qty"), default=1)

    # DB1 is the master registry, not scan history. Keep runtime scan evidence
    # in DB2 only, even if the workbook has old scan columns.
    record["rssi_dbm"] = None
    record["signal_quality"] = None
    record["scan_date"] = None
    record["scan_time"] = None
    record["last_scanned_iso"] = None

    return record


def _validate_record(record: dict[str, Any], row_number: int) -> str | None:
    for field_name in ("record_id", "asset_name", "epc", "tag_id"):
        if not record.get(field_name):
            return f"Missing {field_name} at Excel row {row_number}"
    return None


def _clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text or None
    return value


def _to_int(value: Any, *, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default
