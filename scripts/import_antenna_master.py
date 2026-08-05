import sqlite3
from pathlib import Path
from openpyxl import load_workbook

DB_PATH = Path("rfid_lab.db")


def clean(v):
    if v is None:
        return ""
    return str(v).strip()


def norm_header(v):
    return clean(v).lower().replace(" ", "_").replace("(", "").replace(")", "").replace("/", "_")


def sheet_rows(ws):
    headers = [norm_header(c.value) for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() for v in row):
            continue
        rows.append({headers[i]: row[i] for i in range(len(headers))})
    return rows


def get(row, *names, default=""):
    for n in names:
        key = norm_header(n)
        if key in row and row[key] not in (None, ""):
            return row[key]
    return default


def create_tables(conn):
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS rfid_zones (
        zone_id TEXT PRIMARY KEY,
        zone_name TEXT,
        mac_address TEXT,
        reader TEXT,
        origin_x_m REAL,
        origin_y_m REAL,
        width_m REAL,
        height_m REAL,
        area_m2 REAL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS rfid_readers (
        reader_id TEXT PRIMARY KEY,
        reader_name TEXT,
        mac_pool TEXT,
        zones_served TEXT,
        antenna_count INTEGER,
        tx_power TEXT,
        frequency TEXT,
        protocol TEXT,
        ip_address TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS rfid_antennas (
        antenna_id TEXT PRIMARY KEY,
        mac_address TEXT,
        reader TEXT,
        reader_id TEXT,
        zone_id TEXT,
        port INTEGER,
        position_x_m REAL,
        position_y_m REAL,
        tx_power_dbm REAL,
        polarization TEXT,
        ip_address TEXT
    )
    """)

    conn.commit()


def import_zones(conn, wb):
    if "Zones" not in wb.sheetnames:
        return 0

    rows = sheet_rows(wb["Zones"])
    cur = conn.cursor()
    count = 0

    for r in rows:
        zone_id = clean(get(r, "Zone ID", "Zone"))
        if not zone_id:
            continue

        cur.execute("""
        INSERT INTO rfid_zones (
            zone_id, zone_name, mac_address, reader,
            origin_x_m, origin_y_m, width_m, height_m, area_m2
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(zone_id) DO UPDATE SET
            zone_name=excluded.zone_name,
            mac_address=excluded.mac_address,
            reader=excluded.reader,
            origin_x_m=excluded.origin_x_m,
            origin_y_m=excluded.origin_y_m,
            width_m=excluded.width_m,
            height_m=excluded.height_m,
            area_m2=excluded.area_m2
        """, (
            zone_id,
            clean(get(r, "Name", "Zone Name")),
            clean(get(r, "MAC Address", "MAC")),
            clean(get(r, "Reader")),
            get(r, "Origin X (m)", "Origin X", default=None),
            get(r, "Origin Y (m)", "Origin Y", default=None),
            get(r, "Width (m)", "Width", default=None),
            get(r, "Height (m)", "Height", default=None),
            get(r, "Area (m²)", "Area", default=None),
        ))
        count += 1

    conn.commit()
    return count


def import_readers(conn, wb):
    if "Readers" not in wb.sheetnames:
        return 0

    rows = sheet_rows(wb["Readers"])
    cur = conn.cursor()
    count = 0

    for r in rows:
        reader_id = clean(get(r, "Reader ID", "Reader"))
        if not reader_id:
            continue

        cur.execute("""
        INSERT INTO rfid_readers (
            reader_id, reader_name, mac_pool, zones_served,
            antenna_count, tx_power, frequency, protocol, ip_address
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(reader_id) DO UPDATE SET
            reader_name=excluded.reader_name,
            mac_pool=excluded.mac_pool,
            zones_served=excluded.zones_served,
            antenna_count=excluded.antenna_count,
            tx_power=excluded.tx_power,
            frequency=excluded.frequency,
            protocol=excluded.protocol,
            ip_address=excluded.ip_address
        """, (
            reader_id,
            clean(get(r, "Name", "Reader Name")),
            clean(get(r, "MAC Pool", "MAC Address", "MAC")),
            clean(get(r, "Zones Served", "Zones")),
            get(r, "Antenna Count", default=None),
            clean(get(r, "TX Power")),
            clean(get(r, "Frequency")),
            clean(get(r, "Protocol")),
            clean(get(r, "IP Address", "Reader IP", "IP")),
        ))
        count += 1

    conn.commit()
    return count


def import_antennas(conn, wb):
    if "Antennas" not in wb.sheetnames:
        return 0

    rows = sheet_rows(wb["Antennas"])
    cur = conn.cursor()
    count = 0

    for r in rows:
        antenna_id = clean(get(r, "Antenna ID", "Antenna"))
        if not antenna_id:
            continue

        reader_name = clean(get(r, "Reader"))
        zone_id = clean(get(r, "Zone", "Zone ID"))

        cur.execute("""
        INSERT INTO rfid_antennas (
            antenna_id, mac_address, reader, reader_id, zone_id, port,
            position_x_m, position_y_m, tx_power_dbm, polarization, ip_address
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(antenna_id) DO UPDATE SET
            mac_address=excluded.mac_address,
            reader=excluded.reader,
            reader_id=excluded.reader_id,
            zone_id=excluded.zone_id,
            port=excluded.port,
            position_x_m=excluded.position_x_m,
            position_y_m=excluded.position_y_m,
            tx_power_dbm=excluded.tx_power_dbm,
            polarization=excluded.polarization,
            ip_address=excluded.ip_address
        """, (
            antenna_id,
            clean(get(r, "MAC Address", "MAC")),
            reader_name,
            reader_name.replace("Radar ", "R") if reader_name.startswith("Radar ") else reader_name,
            zone_id,
            get(r, "Port", default=None),
            get(r, "Position X (m)", "Position X", default=None),
            get(r, "Position Y (m)", "Position Y", default=None),
            get(r, "TX Power (dBm)", "TX Power", default=None),
            clean(get(r, "Polarization")),
            clean(get(r, "IP Address", "Antenna IP", "Reader IP", "IP")),
        ))
        count += 1

    conn.commit()
    return count


def main():
    xlsx = Path("data/rfid_master_data.xlsx")
    if not xlsx.exists():
        raise SystemExit("Missing file: data/rfid_master_data.xlsx")

    wb = load_workbook(xlsx, data_only=True)

    conn = sqlite3.connect(DB_PATH)
    create_tables(conn)

    z = import_zones(conn, wb)
    r = import_readers(conn, wb)
    a = import_antennas(conn, wb)

    conn.close()

    print("✅ RFID master antenna data imported")
    print(f"Zones imported: {z}")
    print(f"Readers imported: {r}")
    print(f"Antennas imported: {a}")
    print("Tables created/updated: rfid_zones, rfid_readers, rfid_antennas")


if __name__ == "__main__":
    main()
