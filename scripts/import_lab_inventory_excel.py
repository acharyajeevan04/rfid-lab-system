#!/usr/bin/env python3
"""
Import lab RFID inventory workbook into the current SEARLab RFID SQLite database.

What it does:
1. Loads DB1_MASTER from RFID_System_FINAL_Lab_2026.xlsx into db1_master.
2. Uses Tag_ID as SKU, and BARCODE_VERIFICATION sheet to fill GTIN/barcode where available.
3. Imports optional DB2_INCOMING sample rows and SCAN_SESSIONS rows from the workbook.
4. Backfills every existing DB2 scan so EPCs show item names, SKU, GTIN, and MATCHED/MISMATCH/UNKNOWN status.
5. Exports unmapped EPC review CSV for professor/database cleanup.

Run from project root:
  python scripts/import_lab_inventory_excel.py data/RFID_System_FINAL_Lab_2026.xlsx
"""

from __future__ import annotations

import argparse
import csv
import re
import sqlite3
import sys
from datetime import datetime, date, time
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dependency: openpyxl")
    print("Run: python -m pip install openpyxl")
    sys.exit(1)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DB = PROJECT_ROOT / "rfid_lab.db"

DB1_COLUMNS = {
    "record_id": "TEXT UNIQUE",
    "asset_name": "TEXT",
    "epc": "TEXT",
    "tag_id": "TEXT",
    "component_location": "TEXT",
    "zone_code": "TEXT",
    "zone_name": "TEXT",
    "sub_zone": "TEXT",
    "category": "TEXT",
    "manufacturer": "TEXT",
    "model": "TEXT",
    "unit": "TEXT",
    "qty": "INTEGER DEFAULT 1",
    "status": "TEXT",
    "reader_id": "TEXT",
    "reader_type": "TEXT",
    "verification_method": "TEXT",
    "manual_check": "TEXT",
    "rssi_dbm": "INTEGER",
    "signal_quality": "TEXT",
    "scan_date": "TEXT",
    "scan_time": "TEXT",
    "last_scanned_iso": "TEXT",
    "sku": "TEXT",
    "gtin": "TEXT",
    "tag_source": "TEXT",
    "expected_antenna_id": "INTEGER",
    "item_type": "TEXT",
    "comments": "TEXT",
    "created_at": "TEXT",
    "updated_at": "TEXT",
}

DB2_COLUMNS = {
    "scan_id": "TEXT UNIQUE",
    "epc": "TEXT",
    "scanned_zone": "TEXT",
    "scan_date": "TEXT",
    "scan_time": "TEXT",
    "rssi_dbm": "INTEGER",
    "antenna_id": "INTEGER",
    "read_count": "INTEGER DEFAULT 1",
    "assigned_zone": "TEXT",
    "zone_confidence": "REAL",
    "zone_reason": "TEXT",
    "sku": "TEXT",
    "gtin": "TEXT",
    "tag_classification": "TEXT",
    "reader_id": "TEXT",
    "verification_status": "TEXT DEFAULT 'PENDING'",
    "matched_asset_name": "TEXT",
    "matched_record_id": "TEXT",
    "notes": "TEXT",
    "source": "TEXT DEFAULT 'scanner'",
    "created_at": "TEXT",
}

SESSION_COLUMNS = {
    "session_id": "TEXT UNIQUE",
    "scan_date": "TEXT",
    "scan_time": "TEXT",
    "reader_id": "TEXT",
    "reader_type": "TEXT",
    "zone_covered": "TEXT",
    "unique_tags": "INTEGER DEFAULT 0",
    "total_reads": "INTEGER DEFAULT 0",
    "duration_sec": "INTEGER DEFAULT 0",
    "session_type": "TEXT",
    "notes": "TEXT",
    "created_at": "TEXT",
}

HEADER_MAP_DB1 = {
    "record_id": "record_id",
    "asset_name": "asset_name",
    "epc": "epc",
    "tag_id": "tag_id",
    "component_location": "component_location",
    "zone_code": "zone_code",
    "zone_name": "zone_name",
    "sub_zone": "sub_zone",
    "category": "category",
    "manufacturer": "manufacturer",
    "model": "model",
    "unit": "unit",
    "qty": "qty",
    "status": "status",
    "reader_id": "reader_id",
    "reader_type": "reader_type",
    "verification_method": "verification_method",
    "manual_check": "manual_check",
    "rssi_dbm": "rssi_dbm",
    "signal_quality": "signal_quality",
    "scan_date": "scan_date",
    "scan_time": "scan_time",
    "last_scanned_iso": "last_scanned_iso",
    "comments": "comments",
}

HEADER_MAP_DB2 = {
    "scan_id": "scan_id",
    "epc": "epc",
    "scanned_zone": "scanned_zone",
    "scan_date": "scan_date",
    "scan_time": "scan_time",
    "rssi_dbm": "rssi_dbm",
    "reader_id": "reader_id",
    "verification_status": "verification_status",
    "matched_asset_name": "matched_asset_name",
    "matched_record_id": "matched_record_id",
    "notes": "notes",
}

HEADER_MAP_SESSIONS = {
    "session_id": "session_id",
    "scan_date": "scan_date",
    "scan_time": "scan_time",
    "reader_id": "reader_id",
    "reader_type": "reader_type",
    "zone_covered": "zone_covered",
    "unique_tags": "unique_tags",
    "total_reads": "total_reads",
    "duration_sec": "duration_sec",
    "session_type": "session_type",
    "notes": "notes",
}


def norm_header(v: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(v or "").strip().lower()).strip("_")


def clean_epc(v: Any) -> str:
    return re.sub(r"\s+", "", str(v or "").strip().upper())


def clean_text(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, (datetime, date, time)):
        return format_value(v)
    s = str(v).strip()
    return s if s != "" else None


def format_value(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%dT%H:%M:%SZ")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, time):
        return v.strftime("%H:%M:%S")
    return v


def to_int(v: Any) -> Optional[int]:
    if v in (None, ""):
        return None
    try:
        return int(float(v))
    except Exception:
        return None


def ensure_table(conn: sqlite3.Connection, table: str, columns: Dict[str, str]) -> None:
    cur = conn.cursor()
    cols_sql = ",\n        ".join([f"{name} {typ}" for name, typ in columns.items()])
    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS {table} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            {cols_sql}
        )
    """)
    existing = {row[1] for row in cur.execute(f"PRAGMA table_info({table})").fetchall()}
    for name, typ in columns.items():
        if name not in existing:
            cur.execute(f"ALTER TABLE {table} ADD COLUMN {name} {typ}")
    # Helpful indexes for matching/search
    if table == "db1_master":
        cur.execute("CREATE INDEX IF NOT EXISTS idx_db1_epc ON db1_master(epc)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_db1_sku ON db1_master(sku)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_db1_gtin ON db1_master(gtin)")
    elif table == "db2_incoming":
        cur.execute("CREATE INDEX IF NOT EXISTS idx_db2_epc ON db2_incoming(epc)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_db2_status ON db2_incoming(verification_status)")
    conn.commit()


def rows_from_sheet(ws, header_row: int = 1) -> List[Dict[str, Any]]:
    headers = [norm_header(c.value) for c in ws[header_row]]
    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(v is not None and str(v).strip() != "" for v in row):
            continue
        item = {}
        for h, v in zip(headers, row):
            if h:
                item[h] = format_value(v)
        rows.append(item)
    return rows


def extract_barcode_map(wb) -> Dict[str, Dict[str, str]]:
    out: Dict[str, Dict[str, str]] = {}
    if "BARCODE_VERIFICATION" not in wb.sheetnames:
        return out
    ws = wb["BARCODE_VERIFICATION"]
    rows = rows_from_sheet(ws, header_row=4)
    for r in rows:
        rec_id = clean_text(r.get("db1_record"))
        label = clean_text(r.get("barcode_sku_on_label"))
        if not rec_id or not label:
            continue
        parts = [p.strip() for p in str(label).split("/")]
        numeric = next((p for p in parts if re.fullmatch(r"\d{8,14}", p)), None)
        sku_part = next((p for p in parts if not re.fullmatch(r"\d{8,14}", p)), None)
        out[rec_id] = {
            "barcode_label": label,
            "gtin": numeric or "",
            "sku_from_barcode": sku_part or "",
        }
    return out


def classify_tag(epc: str, exists: bool) -> str:
    if exists:
        return "registered"
    if epc.upper().startswith("E280"):
        return "demo-e280-or-unmapped"
    return "unknown"


def upsert(conn: sqlite3.Connection, table: str, key_col: str, row: Dict[str, Any], allowed_cols: Dict[str, str]) -> str:
    cur = conn.cursor()
    cols = [c for c in allowed_cols.keys() if c in row]
    if key_col not in row or not row.get(key_col):
        raise ValueError(f"Missing key column {key_col}")
    existing = cur.execute(f"SELECT id FROM {table} WHERE UPPER(TRIM({key_col})) = UPPER(TRIM(?))", (str(row[key_col]),)).fetchone()
    now = datetime.utcnow().isoformat()
    if "updated_at" in allowed_cols:
        row["updated_at"] = now
        if "updated_at" not in cols:
            cols.append("updated_at")
    if existing:
        set_cols = [c for c in cols if c != key_col]
        if set_cols:
            sql = f"UPDATE {table} SET " + ", ".join([f"{c}=?" for c in set_cols]) + f" WHERE id=?"
            cur.execute(sql, [row.get(c) for c in set_cols] + [existing[0]])
        return "updated"
    else:
        if "created_at" in allowed_cols:
            row.setdefault("created_at", now)
            if "created_at" not in cols:
                cols.append("created_at")
        placeholders = ",".join(["?"] * len(cols))
        sql = f"INSERT INTO {table} ({','.join(cols)}) VALUES ({placeholders})"
        cur.execute(sql, [row.get(c) for c in cols])
        return "inserted"


def import_db1(conn: sqlite3.Connection, wb) -> Dict[str, int]:
    if "DB1_MASTER" not in wb.sheetnames:
        raise ValueError("Workbook is missing DB1_MASTER sheet")
    ws = wb["DB1_MASTER"]
    rows = rows_from_sheet(ws, header_row=1)
    barcode_map = extract_barcode_map(wb)
    inserted = updated = skipped = 0
    for r in rows:
        mapped: Dict[str, Any] = {}
        for src, dst in HEADER_MAP_DB1.items():
            mapped[dst] = clean_text(r.get(src))
        record_id = clean_text(mapped.get("record_id"))
        epc = clean_epc(mapped.get("epc"))
        if not record_id or not epc:
            skipped += 1
            continue
        mapped["record_id"] = record_id
        mapped["epc"] = epc
        mapped["qty"] = to_int(mapped.get("qty")) or 1
        mapped["rssi_dbm"] = to_int(mapped.get("rssi_dbm"))
        mapped["sku"] = clean_text(mapped.get("tag_id"))
        # GTIN/barcode only when actual physical barcode exists.
        bm = barcode_map.get(record_id, {})
        mapped["gtin"] = bm.get("gtin") or None
        if bm.get("sku_from_barcode") and not mapped["sku"]:
            mapped["sku"] = bm["sku_from_barcode"]
        mapped["tag_source"] = "inventory-workbook"
        mapped["item_type"] = mapped.get("category")
        res = upsert(conn, "db1_master", "epc", mapped, DB1_COLUMNS)
        inserted += res == "inserted"
        updated += res == "updated"
    conn.commit()
    return {"inserted": inserted, "updated": updated, "skipped": skipped}


def import_db2_sample(conn: sqlite3.Connection, wb) -> Dict[str, int]:
    if "DB2_INCOMING" not in wb.sheetnames:
        return {"inserted": 0, "updated": 0, "skipped": 0}
    ws = wb["DB2_INCOMING"]
    rows = rows_from_sheet(ws, header_row=1)
    inserted = updated = skipped = 0
    for r in rows:
        mapped: Dict[str, Any] = {}
        for src, dst in HEADER_MAP_DB2.items():
            mapped[dst] = clean_text(r.get(src))
        if not mapped.get("scan_id") or not mapped.get("epc"):
            skipped += 1
            continue
        mapped["epc"] = clean_epc(mapped["epc"])
        mapped["rssi_dbm"] = to_int(mapped.get("rssi_dbm"))
        mapped["source"] = "inventory_workbook_sample"
        mapped.setdefault("created_at", datetime.utcnow().isoformat())
        res = upsert(conn, "db2_incoming", "scan_id", mapped, DB2_COLUMNS)
        inserted += res == "inserted"
        updated += res == "updated"
    conn.commit()
    return {"inserted": inserted, "updated": updated, "skipped": skipped}


def import_sessions(conn: sqlite3.Connection, wb) -> Dict[str, int]:
    if "SCAN_SESSIONS" not in wb.sheetnames:
        return {"inserted": 0, "updated": 0, "skipped": 0}
    ws = wb["SCAN_SESSIONS"]
    rows = rows_from_sheet(ws, header_row=1)
    inserted = updated = skipped = 0
    for r in rows:
        mapped: Dict[str, Any] = {}
        for src, dst in HEADER_MAP_SESSIONS.items():
            mapped[dst] = clean_text(r.get(src))
        if not mapped.get("session_id"):
            skipped += 1
            continue
        mapped["unique_tags"] = to_int(mapped.get("unique_tags")) or 0
        mapped["total_reads"] = to_int(mapped.get("total_reads")) or 0
        mapped["duration_sec"] = to_int(mapped.get("duration_sec")) or 0
        mapped.setdefault("created_at", datetime.utcnow().isoformat())
        res = upsert(conn, "scan_sessions", "session_id", mapped, SESSION_COLUMNS)
        inserted += res == "inserted"
        updated += res == "updated"
    conn.commit()
    return {"inserted": inserted, "updated": updated, "skipped": skipped}


def backfill_scans(conn: sqlite3.Connection) -> Dict[str, int]:
    cur = conn.cursor()
    asset_rows = cur.execute("SELECT record_id, asset_name, epc, sku, gtin, zone_code, zone_name FROM db1_master WHERE epc IS NOT NULL").fetchall()
    master = {clean_epc(r[2]): r for r in asset_rows if r[2]}
    scans = cur.execute("SELECT id, epc, scanned_zone, verification_status FROM db2_incoming WHERE epc IS NOT NULL").fetchall()
    matched = mismatch = unknown = duplicate_kept = 0
    for scan_id, epc_raw, scanned_zone, old_status in scans:
        epc = clean_epc(epc_raw)
        asset = master.get(epc)
        if asset:
            record_id, asset_name, _, sku, gtin, zone_code, zone_name = asset
            sz = clean_text(scanned_zone)
            mz = clean_text(zone_code)
            if sz and mz and sz.upper() != mz.upper():
                status = "MISMATCH"
                mismatch += 1
                notes = f"Known item: {asset_name} | SKU: {sku or ''} | Expected zone {mz}, scanned zone {sz}"
            else:
                if old_status == "DUPLICATE":
                    status = "DUPLICATE"
                    duplicate_kept += 1
                else:
                    status = "MATCHED"
                    matched += 1
                notes = f"Tag verified — EPC in DB1. Item: {asset_name} | SKU: {sku or ''} | GTIN: {gtin or ''}"
            cur.execute("""
                UPDATE db2_incoming
                SET verification_status=?, matched_asset_name=?, matched_record_id=?, sku=?, gtin=?,
                    tag_classification='registered', notes=?
                WHERE id=?
            """, (status, asset_name, record_id, sku, gtin, notes, scan_id))
        else:
            unknown += 1
            cur.execute("""
                UPDATE db2_incoming
                SET verification_status='UNKNOWN', matched_asset_name='— Not in Master Table —', matched_record_id='—',
                    sku=NULL, gtin=NULL, tag_classification=?,
                    notes='EPC not found in master table. Review with EPC decoder or remove if not attached to a real item.'
                WHERE id=?
            """, (classify_tag(epc, False), scan_id))
    # Update last scanned info in DB1 from latest matching scan
    latest_rows = cur.execute("""
        SELECT epc, scan_date, scan_time, created_at
        FROM db2_incoming
        WHERE epc IS NOT NULL
        ORDER BY id DESC
    """).fetchall()
    seen = set()
    for epc_raw, scan_date, scan_time, created_at in latest_rows:
        epc = clean_epc(epc_raw)
        if not epc or epc in seen:
            continue
        seen.add(epc)
        cur.execute("""
            UPDATE db1_master
            SET scan_date=?, scan_time=?, last_scanned_iso=?, updated_at=?
            WHERE UPPER(TRIM(epc))=UPPER(TRIM(?))
        """, (scan_date, scan_time, created_at or datetime.utcnow().isoformat(), datetime.utcnow().isoformat(), epc))
    conn.commit()
    return {"matched": matched, "mismatch": mismatch, "unknown": unknown, "duplicate_kept": duplicate_kept}


def export_unmapped(conn: sqlite3.Connection, export_dir: Path) -> Path:
    export_dir.mkdir(parents=True, exist_ok=True)
    path = export_dir / "unmapped_tags_review.csv"
    cur = conn.cursor()
    rows = cur.execute("""
        SELECT s.epc, COUNT(*) AS read_events, MIN(s.created_at) AS first_seen, MAX(s.created_at) AS last_seen,
               MAX(s.rssi_dbm) AS strongest_rssi
        FROM db2_incoming s
        WHERE UPPER(TRIM(s.epc)) NOT IN (SELECT UPPER(TRIM(epc)) FROM db1_master WHERE epc IS NOT NULL)
        GROUP BY s.epc
        ORDER BY COUNT(*) DESC
    """).fetchall()
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["epc", "read_events", "first_seen", "last_seen", "strongest_rssi", "likely_demo_e280", "recommendation"])
        for epc, read_events, first_seen, last_seen, strongest_rssi in rows:
            epc_clean = clean_epc(epc)
            likely_e280 = epc_clean.startswith("E280")
            rec = "Likely local/demo E280 tag; verify physically, then destroy/remove if not attached to a real item" if likely_e280 else "Review with EPC decoder and map to SKU/GTIN if real"
            writer.writerow([epc_clean, read_events, first_seen, last_seen, strongest_rssi, likely_e280, rec])
    return path


def summary(conn: sqlite3.Connection) -> Dict[str, Any]:
    cur = conn.cursor()
    out = {}
    out["db1_master_unique_items"] = cur.execute("SELECT COUNT(*) FROM db1_master").fetchone()[0]
    out["db1_unique_epcs"] = cur.execute("SELECT COUNT(DISTINCT UPPER(TRIM(epc))) FROM db1_master WHERE epc IS NOT NULL").fetchone()[0]
    out["db2_scan_events"] = cur.execute("SELECT COUNT(*) FROM db2_incoming").fetchone()[0]
    status_rows = cur.execute("SELECT verification_status, COUNT(*) FROM db2_incoming GROUP BY verification_status").fetchall()
    out["db2_status_breakdown"] = {k or "NULL": v for k, v in status_rows}
    out["unmapped_unique_tags"] = cur.execute("""
        SELECT COUNT(DISTINCT UPPER(TRIM(epc))) FROM db2_incoming
        WHERE UPPER(TRIM(epc)) NOT IN (SELECT UPPER(TRIM(epc)) FROM db1_master WHERE epc IS NOT NULL)
    """).fetchone()[0]
    return out


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", nargs="?", default=str(PROJECT_ROOT / "data" / "RFID_System_FINAL_Lab_2026.xlsx"), help="Path to RFID_System_FINAL_Lab_2026.xlsx")
    parser.add_argument("--db", default=str(DEFAULT_DB), help="Path to rfid_lab.db")
    parser.add_argument("--no-sample-scans", action="store_true", help="Do not import DB2_INCOMING sample rows from workbook")
    parser.add_argument("--no-sessions", action="store_true", help="Do not import SCAN_SESSIONS rows from workbook")
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser().resolve()
    db_path = Path(args.db).expanduser().resolve()

    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")
    if not db_path.exists():
        print(f"Database does not exist yet. It will be created: {db_path}")

    print(f"Reading workbook: {workbook_path}")
    print(f"Using database:   {db_path}")

    wb = load_workbook(workbook_path, data_only=True)
    conn = sqlite3.connect(db_path)
    try:
        ensure_table(conn, "db1_master", DB1_COLUMNS)
        ensure_table(conn, "db2_incoming", DB2_COLUMNS)
        ensure_table(conn, "scan_sessions", SESSION_COLUMNS)

        db1_result = import_db1(conn, wb)
        print(f"DB1 master import: {db1_result}")

        if not args.no_sample_scans:
            db2_result = import_db2_sample(conn, wb)
            print(f"DB2 sample scan import: {db2_result}")

        if not args.no_sessions:
            sess_result = import_sessions(conn, wb)
            print(f"Scan sessions import: {sess_result}")

        backfill_result = backfill_scans(conn)
        print(f"DB2 backfill/matching: {backfill_result}")

        unmapped_path = export_unmapped(conn, PROJECT_ROOT / "exports")
        print(f"Unmapped tag review exported: {unmapped_path}")

        print("\nFINAL DATABASE SUMMARY")
        for k, v in summary(conn).items():
            print(f"  {k}: {v}")

        print("\nDONE. Existing UI will now show item names/SKU/GTIN for matching EPCs.")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
